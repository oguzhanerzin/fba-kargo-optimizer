import os
import io
import json
import traceback
from flask import Flask, request, jsonify, send_file, render_template
import openpyxl
from optimizer import optimize
from excel_builder import build_excel

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10 MB

def parse_num(v):
    try: return float(str(v).replace(',', '.'))
    except: return 0.0

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/optimize', methods=['POST'])
def run_optimize():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Dosya bulunamadı'}), 400

        f = request.files['file']
        if not f.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Sadece .xlsx veya .xls dosyaları desteklenir'}), 400

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)

        # Veri sayfasını bul
        ws = None
        for name in wb.sheetnames:
            if 'veri' in name.lower() or 'data' in name.lower() or name == wb.sheetnames[0]:
                ws = wb[name]
                break

        if ws is None:
            return jsonify({'error': 'Veri sayfası bulunamadı'}), 400

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({'error': 'Veri sayfası boş'}), 400

        # Başlık satırını tespit et
        baslik_satir = None
        for i, row in enumerate(rows[:5]):
            row_str = [str(c or '').lower() for c in row]
            if any('asin' in c for c in row_str):
                baslik_satir = i
                break

        if baslik_satir is None:
            return jsonify({'error': 'ASIN sütunu bulunamadı. Başlık satırında "ASIN" geçmeli.'}), 400

        # Sütun indekslerini bul
        baslik = [str(c or '').lower() for c in rows[baslik_satir]]
        def col_idx(keywords):
            for kw in keywords:
                for i, h in enumerate(baslik):
                    if kw in h: return i
            return None

        i_asin    = col_idx(['asin'])
        i_ad      = col_idx(['ürün', 'ad', 'sku', 'isim', 'name'])
        i_gercek  = col_idx(['ağırlık', 'agirlik', 'weight', 'lb', 'birim a'])
        i_hacimsel= col_idx(['hacimsel', 'volumetric', 'lbs', 'birim h'])
        i_adet    = col_idx(['adet', 'quantity', 'qty', 'miktar'])

        if i_asin is None:
            return jsonify({'error': 'ASIN sütunu bulunamadı'}), 400
        if i_gercek is None:
            return jsonify({'error': 'Ağırlık sütunu bulunamadı'}), 400
        if i_adet is None:
            return jsonify({'error': 'Adet sütunu bulunamadı'}), 400

        # Veriyi oku
        sku_listesi = []
        for row in rows[baslik_satir + 1:]:
            asin  = str(row[i_asin] or '').strip() if i_asin < len(row) else ''
            ad    = str(row[i_ad] or '').strip()   if i_ad is not None and i_ad < len(row) else ''
            g     = parse_num(row[i_gercek])        if i_gercek < len(row) else 0
            h     = parse_num(row[i_hacimsel])      if i_hacimsel is not None and i_hacimsel < len(row) else 0
            adet  = int(parse_num(row[i_adet]))     if i_adet < len(row) else 0

            if not asin or adet <= 0 or g <= 0:
                continue

            sku_listesi.append({'asin': asin, 'ad': ad, 'gercek': g, 'hacimsel': h, 'adet': adet})

        if not sku_listesi:
            return jsonify({'error': 'Geçerli ürün satırı bulunamadı'}), 400

        # Optimizasyonu çalıştır
        result = optimize(sku_listesi)

        if 'error' in result:
            return jsonify({'error': result['error']}), 500

        # Excel dosyasını oluştur
        excel_bytes = build_excel(result)

        return send_file(
            io.BytesIO(excel_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='kargo_plani_optimized.xlsx'
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Sunucu hatası: {str(e)}'}), 500

@app.route('/preview', methods=['POST'])
def preview():
    """Sadece istatistik döner, Excel indirmez — UI önizleme için"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Dosya bulunamadı'}), 400

        f = request.files['file']
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        baslik_satir = None
        for i, row in enumerate(rows[:5]):
            row_str = [str(c or '').lower() for c in row]
            if any('asin' in c for c in row_str):
                baslik_satir = i; break

        if baslik_satir is None:
            # İlk satır veri kabul et
            baslik_satir = -1

        baslik = [str(c or '').lower() for c in rows[max(0, baslik_satir)]]
        def col_idx(keywords):
            for kw in keywords:
                for i, h in enumerate(baslik):
                    if kw in h: return i
            return None

        i_asin   = col_idx(['asin'])
        i_gercek = col_idx(['ağırlık','agirlik','weight','lb','birim a'])
        i_adet   = col_idx(['adet','quantity','qty','miktar'])

        sku_listesi = []
        for row in rows[baslik_satir + 1:]:
            try:
                asin = str(row[i_asin] or '').strip() if i_asin is not None else ''
                g    = float(str(row[i_gercek] or 0).replace(',','.')) if i_gercek is not None else 0
                adet = int(float(str(row[i_adet] or 0).replace(',','.'))) if i_adet is not None else 0
                if asin and adet > 0 and g > 0:
                    sku_listesi.append({'asin': asin, 'gercek': g, 'adet': adet})
            except: continue

        import math
        toplam_g = sum(s['gercek'] * s['adet'] for s in sku_listesi)
        return jsonify({
            'sku_sayisi'     : len(sku_listesi),
            'toplam_adet'    : sum(s['adet'] for s in sku_listesi),
            'toplam_gercek'  : round(toplam_g, 2),
            'tahmini_koli'   : math.ceil(toplam_g / 46) if toplam_g > 0 else 0,
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'solver': 'OR-Tools CP-SAT'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
