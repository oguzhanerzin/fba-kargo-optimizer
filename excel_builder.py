"""
Optimizasyon sonucundan Excel dosyası üret
4 sayfa: Veri, Koli_Ozet, Koli_Detay, Box_Product_Summary
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Renkler ──────────────────────────────────────────────────────────────────
C = {
    'lacivert' : 'FF1F3864',
    'mavi'     : 'FF2E75B6',
    'acik_mavi': 'FFD6E4F0',
    'alt_satir': 'FFEBF3FB',
    'yesil'    : 'FFE2EFDA',
    'sari'     : 'FFFFF2CC',
    'kirmizi'  : 'FFFCE4D6',
    'gri'      : 'FFF2F2F2',
    'beyaz'    : 'FFFFFFFF',
    'mor'      : 'FFF3E5F5',
    'turuncu'  : 'FFFFF0E0',
}

KOLI_RENKLER = ['FFD6E4F0','FFE2EFDA','FFFFF2CC','FFF3E5F5','FFFFF0E0',
                'FFE8F5E9','FFFFF3E0','FFE0F7FA','FFFBE9E7','FFE8EAF6']

MAKS_GERCEK    = 46.0
MAKS_HACIMSEL  = 64.0
IDEAL_HACIMSEL = 49.0


def _fill(hex6): return PatternFill('solid', fgColor=hex6)
def _font(bold=False, size=10, color='FF000000', name='Calibri'):
    return Font(name=name, bold=bold, size=size, color=color)
def _center(wrap=False): return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def _left():             return Alignment(horizontal='left',   vertical='center')
def _border():
    t = Side(style='thin', color='FFB0C4DE')
    return Border(left=t, right=t, top=t, bottom=t)

def _hdr(ws, row, col, cols_merge, value, bg, fg='FFFFFFFF', size=10, bold=True, wrap=False):
    if cols_merge > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + cols_merge - 1)
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=bold, size=size, color=fg)
    cell.alignment = _center(wrap=wrap)
    cell.border    = _border()
    return cell

def _cell(ws, row, col, value, bg='FFFFFFFF', bold=False, align='center', fmt=None, size=9):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=bold, size=size)
    cell.alignment = _center() if align == 'center' else _left()
    cell.border    = _border()
    if fmt: cell.number_format = fmt
    return cell


def build_excel(result: dict) -> bytes:
    koliler = result['koliler']
    stats   = result['stats']
    skular  = result['skular']

    wb = Workbook()
    wb.remove(wb.active)

    _veri_sayfasi(wb, skular)
    _ozet_sayfasi(wb, koliler, stats)
    _detay_sayfasi(wb, koliler)
    _pivot_sayfasi(wb, koliler, skular)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 1. VERİ SAYFASI ───────────────────────────────────────────────────────────
def _veri_sayfasi(wb, skular):
    ws = wb.create_sheet('📦 Veri')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 26

    _hdr(ws, 1, 1, 5, '📦  VERİ GİRİŞİ — Ürün Listesi', C['lacivert'], size=14)
    for c, h in enumerate(['ASIN', 'Ürün Adı', 'Birim Ağırlık (lb)',
                            'Birim Hacimsel Ağ. (lbs)', 'Adet'], 1):
        _hdr(ws, 2, c, 1, h, C['mavi'], size=10)

    for i, sku in enumerate(skular):
        r   = i + 3
        bg  = C['beyaz'] if i % 2 == 0 else C['alt_satir']
        _cell(ws, r, 1, sku['asin'],    bg, align='left', size=9)
        _cell(ws, r, 2, sku['ad'],      bg, align='left', size=9)
        _cell(ws, r, 3, sku['gercek'],  bg, fmt='0.00', size=9)
        _cell(ws, r, 4, sku['hacimsel'],bg, fmt='0.00', size=9)
        _cell(ws, r, 5, sku['adet'],    bg, fmt='0', size=9)
        ws.row_dimensions[r].height = 18

    for col, w in zip(range(1, 6), [140, 160, 155, 175, 60]):
        ws.column_dimensions[get_column_letter(col)].width = w / 7


# ── 2. KOLİ ÖZET SAYFASI ─────────────────────────────────────────────────────
def _ozet_sayfasi(wb, koliler, stats):
    ws = wb.create_sheet('📋 Koli_Ozet')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A5'

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 26

    _hdr(ws, 1, 1, 7, '📋  KOLİ ÖZETİ', C['lacivert'], size=14)

    info = (f'Koli: 21×18×16 inc  |  Maks. Gerçek: {MAKS_GERCEK} lb  |  '
            f'Hacimsel Hedef: {IDEAL_HACIMSEL} lbs  |  Tavan: {MAKS_HACIMSEL} lbs  |  '
            f'Solver: OR-Tools CP-SAT ✅')
    _hdr(ws, 2, 1, 7, info, C['acik_mavi'], fg='FF1F3864', size=9, bold=False, wrap=True)

    dup = stats.get('duplicate_asinler', [])
    dup_msg = f'⚠️ {len(dup)} duplicate ASIN birleştirildi: {", ".join(dup[:5])}{"..." if len(dup)>5 else ""}' if dup else '✅ Duplicate ASIN yok'
    _hdr(ws, 3, 1, 7, dup_msg, C['sari'] if dup else C['yesil'],
         fg='FF7F6000' if dup else 'FF375623', size=9, bold=bool(dup))

    for c, h in enumerate(['Koli', 'Ürün Adedi', 'SKU Sayısı',
                            'Toplam Ağırlık (lb)', 'Toplam Hacim (lbs)',
                            'Boş Ağırlık (lb)', 'Boş Hacim (lbs)'], 1):
        _hdr(ws, 4, c, 1, h, C['lacivert'], size=10)

    for i, koli in enumerate(koliler):
        r   = 5 + i
        bg  = C['beyaz'] if i % 2 == 0 else C['alt_satir']
        bos_g = round(MAKS_GERCEK - koli['toplam_gercek'], 2)
        bos_h = round(IDEAL_HACIMSEL - koli['toplam_hacimsel'], 2)
        h_bg  = (C['kirmizi'] if koli['toplam_hacimsel'] > MAKS_HACIMSEL
                 else C['sari'] if koli['toplam_hacimsel'] > IDEAL_HACIMSEL
                 else C['yesil'])

        _cell(ws, r, 1, koli['no'],                       C['mavi'],  bold=True, fmt='0')
        _cell(ws, r, 2, koli['adet_toplam'],              bg,  fmt='0')
        _cell(ws, r, 3, koli['sku_sayisi'],               bg,  bold=True, fmt='0')
        _cell(ws, r, 4, round(koli['toplam_gercek'],  2), bg,  fmt='0.00')
        _cell(ws, r, 5, round(koli['toplam_hacimsel'],2), h_bg,fmt='0.00')
        _cell(ws, r, 6, bos_g,                            bg,  fmt='0.00')
        _cell(ws, r, 7, bos_h,                            bg,  fmt='0.00')
        ws.row_dimensions[r].height = 22

    # Toplam satırı
    tr = 5 + len(koliler)
    _cell(ws, tr, 1, 'TOPLAM', C['lacivert'], bold=True, size=10)
    ws.cell(tr, 1).font  = _font(bold=True, size=10, color='FFFFFFFF')
    ws.cell(tr, 1).fill  = _fill(C['lacivert'])
    _cell(ws, tr, 2, stats['toplam_adet'], C['lacivert'], bold=True, fmt='0')
    ws.cell(tr, 2).font = _font(bold=True, size=10, color='FFFFFFFF')
    ws.cell(tr, 2).fill = _fill(C['lacivert'])
    for c in range(3, 8):
        ws.cell(tr, c).fill = _fill(C['lacivert'])
    ws.row_dimensions[tr].height = 24

    # Özet istatistik kutusu (sağ taraf, sütun 9-10)
    ozet_data = [
        ('📊 OPTİMİZASYON ÖZETİ', None),
        ('Solver',              'OR-Tools CP-SAT'),
        ('Toplam Koli',         stats['toplam_koli']),
        ('Toplam Ürün Adedi',   stats['toplam_adet']),
        ('SKU/Koli (min–maks)', f"{stats['sku_min']}–{stats['sku_max']}"),
        ('SKU Farkı',           stats['sku_fark']),
        ('Gerçek Ağırlık (min–maks)', f"{stats['gercek_min']}–{stats['gercek_max']} lb"),
        ('Hacimsel (min–maks)', f"{stats['hacimsel_min']}–{stats['hacimsel_max']} lbs"),
        ('Adet Farkı',          stats['adet_fark']),
    ]
    for i, (etiket, deger) in enumerate(ozet_data):
        r = i + 1
        if deger is None:
            ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
            _hdr(ws, r, 9, 1, etiket, C['lacivert'], size=10)
        else:
            _cell(ws, r, 9,  etiket, C['acik_mavi'], bold=True, align='left', size=9)
            _cell(ws, r, 10, deger,  C['beyaz'] if i%2==0 else C['gri'], size=9)
        ws.row_dimensions[r].height = 22

    # Renk açıklaması
    ws.cell(len(ozet_data)+2, 9, '🟢 Hacimsel < 49 lbs').fill = _fill(C['yesil'])
    ws.cell(len(ozet_data)+3, 9, '🟡 Hacimsel 49–64 lbs').fill = _fill(C['sari'])
    ws.cell(len(ozet_data)+4, 9, '🔴 Hacimsel > 64 lbs').fill  = _fill(C['kirmizi'])
    for r in range(len(ozet_data)+2, len(ozet_data)+5):
        ws.cell(r, 9).font = _font(size=9)
        ws.row_dimensions[r].height = 18

    for col, w in zip(range(1, 11), [60,80,80,150,150,130,130,20,160,110]):
        ws.column_dimensions[get_column_letter(col)].width = w / 7


# ── 3. KOLİ DETAY SAYFASI ────────────────────────────────────────────────────
def _detay_sayfasi(wb, koliler):
    ws = wb.create_sheet('🗂 Koli_Detay')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 26
    _hdr(ws, 1, 1, 5, '🗂  KOLİ DETAY — Ürün Bazında Koli Dağılımı', C['lacivert'], size=14)
    for c, h in enumerate(['ASIN','Ürün Adı','Ağırlık (lb)','Hacim (lbs)','Koli'], 1):
        _hdr(ws, 2, c, 1, h, C['lacivert'], size=10)

    satir = 3
    for ki, koli in enumerate(koliler):
        bg_koli = KOLI_RENKLER[ki % len(KOLI_RENKLER)]
        for pi, p in enumerate(koli['parcalar']):
            # Her adet ayrı satır
            for _ in range(p['adet']):
                bg = bg_koli if pi % 2 == 0 else _lighten(bg_koli)
                _cell(ws, satir, 1, p['asin'],    bg, align='left', size=9)
                _cell(ws, satir, 2, p['ad'],      bg, align='left', size=9)
                _cell(ws, satir, 3, p['gercek'],  bg, fmt='0.00',   size=9)
                _cell(ws, satir, 4, p['hacimsel'],bg, fmt='0.00',   size=9)
                _cell(ws, satir, 5, koli['no'],   bg, bold=True,    size=9)
                ws.row_dimensions[satir].height = 18
                satir += 1

    for col, w in zip(range(1, 6), [140,160,110,110,60]):
        ws.column_dimensions[get_column_letter(col)].width = w / 7


# ── 4. PİVOT SAYFASI ─────────────────────────────────────────────────────────
def _pivot_sayfasi(wb, koliler, skular):
    ws = wb.create_sheet('📊 Box_Product_Summary')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'C3'

    K = len(koliler)
    genislik = 2 + K

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 26
    _hdr(ws, 1, 1, genislik, '📊  BOX PRODUCT SUMMARY — SKU × Koli Dağılımı', C['lacivert'], size=14)
    for c, h in enumerate(['ASIN','Ürün Adı'] + [f'Koli {k+1}' for k in range(K)], 1):
        _hdr(ws, 2, c, 1, h, C['lacivert'], size=10)

    # ASIN → koli adetleri
    asin_map = {s['asin']: {'ad': s['ad'], 'adetler': [0]*K} for s in skular}
    for ki, koli in enumerate(koliler):
        for p in koli['parcalar']:
            asin_map[p['asin']]['adetler'][ki] += p['adet']

    for i, (asin, info) in enumerate(sorted(asin_map.items())):
        r  = i + 3
        bg = C['beyaz'] if i % 2 == 0 else C['alt_satir']
        _cell(ws, r, 1, asin,     bg, align='left', size=9)
        _cell(ws, r, 2, info['ad'],bg, align='left', size=9)
        for ki, adet in enumerate(info['adetler']):
            if adet > 0:
                _cell(ws, r, 3+ki, adet, bg, bold=True,  size=9, fmt='0')
                ws.cell(r, 3+ki).font = _font(bold=True, size=9, color=C['lacivert'].replace('FF',''))
            else:
                _cell(ws, r, 3+ki, None, bg, size=9)
        ws.row_dimensions[r].height = 18

    # Toplam satırı
    tr = 3 + len(asin_map)
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    _hdr(ws, tr, 1, 1, 'TOPLAM ADET', C['lacivert'], size=10)
    for ki, koli in enumerate(koliler):
        _cell(ws, tr, 3+ki, koli['adet_toplam'], C['lacivert'], bold=True, size=10)
        ws.cell(tr, 3+ki).font = _font(bold=True, size=10, color='FFFFFFFF')
        ws.cell(tr, 3+ki).fill = _fill(C['lacivert'])
    ws.row_dimensions[tr].height = 24

    ws.column_dimensions['A'].width = 140/7
    ws.column_dimensions['B'].width = 160/7
    for ki in range(K):
        ws.column_dimensions[get_column_letter(3+ki)].width = 65/7


def _lighten(hex8: str) -> str:
    """8 karakterlik hex rengi %15 açıklaştır"""
    prefix = hex8[:2]
    r = min(255, int(hex8[2:4], 16) + 20)
    g = min(255, int(hex8[4:6], 16) + 20)
    b = min(255, int(hex8[6:8], 16) + 20)
    return f'{prefix}{r:02X}{g:02X}{b:02X}'
